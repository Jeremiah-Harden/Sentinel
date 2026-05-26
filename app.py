"""
Jerah's Job Scout — Click a board, pick your roles, get an Excel.
Run: streamlit run app.py
"""

import io
import os
import re
import time
from contextlib import contextmanager
from urllib.parse import urlparse, urljoin, urlencode, parse_qs, urlunparse

import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

# ── Constants ──────────────────────────────────────────────────────────────────

JOB_BOARDS = {
    "DailyRemote":  "https://dailyremote.com/",
    "Wellfound":    "https://wellfound.com/jobs",
    "Dice":         "https://www.dice.com/jobs",
    "USAJobs":      "https://www.usajobs.gov/Search/",
    "RemoteOK":     "https://remoteok.com",
    "LinkedIn":     "https://www.linkedin.com/jobs/search/",
}

QUICK_TERMS = [
    "Cybersecurity Intern",
    "SOC Analyst",
    "Incident Response",
    "Penetration Tester",
    "Cloud Security",
    "GRC Analyst",
    "Software Engineer Intern",
    "Data Analyst",
    "IT Support",
    "Network Engineer",
]

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

COLUMNS = [
    "Job Title", "Company", "Salary", "Location", "Job Type",
    "Date Posted", "Tags", "Description", "Requirements", "Apply URL", "Listing URL",
]


# ── CSS ────────────────────────────────────────────────────────────────────────
# config.toml handles the base theme; this block handles things config.toml can't.

CSS = """
<style>
#MainMenu, footer, header { visibility: hidden; }

/* ── Buttons ── */
div.stButton > button {
    background-color: #0c1e31;
    border: 1.5px solid #1e3a5f;
    color: #8aaac4;
    border-radius: 10px;
    font-size: 13px;
    font-weight: 500;
    transition: border-color 0.15s, color 0.15s, background-color 0.15s;
}
div.stButton > button:hover {
    border-color: #D4AF37;
    color: #f0cc55;
    background-color: rgba(212,175,55,0.08);
}
div.stButton > button:focus:not(:active) {
    border-color: #D4AF37;
    color: #D4AF37;
    box-shadow: 0 0 0 2px rgba(212,175,55,0.25);
}

/* ── Progress bar fill ── */
div[data-testid="stProgressBar"] > div > div {
    background-color: #D4AF37 !important;
}

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #07111f; }
::-webkit-scrollbar-thumb { background: #1e3a5f; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #D4AF37; }
</style>
"""


# ── UI helpers ─────────────────────────────────────────────────────────────────

def step_label(number: int, text: str):
    st.markdown(
        f'<div style="display:flex; align-items:center; gap:10px; margin-bottom:10px;">'
        f'<div style="background:#D4AF37; color:#07111f; width:24px; height:24px; '
        f'border-radius:50%; display:flex; align-items:center; justify-content:center; '
        f'font-size:12px; font-weight:800; flex-shrink:0;">{number}</div>'
        f'<span style="color:#d8e4ef; font-size:14px; font-weight:600;">{text}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )


def phase_label(text: str):
    st.markdown(
        f'<p style="color:#D4AF37; font-size:11px; font-weight:800; '
        f'letter-spacing:2px; text-transform:uppercase; margin:16px 0 8px;">{text}</p>',
        unsafe_allow_html=True,
    )


def status_line(html_body: str, placeholder):
    placeholder.markdown(
        f'<p style="color:#8aaac4; font-size:13px; margin:4px 0;">{html_body}</p>',
        unsafe_allow_html=True,
    )


def gold(text: str) -> str:
    return f'<span style="color:#D4AF37; font-weight:700;">{text}</span>'


def white(text: str) -> str:
    return f'<span style="color:#ffffff;">{text}</span>'


# ── URL helpers ────────────────────────────────────────────────────────────────

def build_url(base_url: str, search_term: str, page: int) -> str:
    parsed = urlparse(base_url)
    params = {k: v[0] for k, v in parse_qs(parsed.query).items()}
    if search_term:
        params["search"] = search_term
    params["page"] = str(page)
    return urlunparse(parsed._replace(query=urlencode(params)))


def salary_type(value: str) -> str:
    v = (value or "").strip().lower()
    if not v or v in ("n/a", "not specified", "not disclosed", ""):
        return "none"
    if re.search(r"[\d$£€]", v):
        return "disclosed"
    return "vague"


# ── Scraping ───────────────────────────────────────────────────────────────────

@contextmanager
def browser_session():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            yield browser
        finally:
            browser.close()


def fetch_html(browser, url: str, wait_ms: int = 2500) -> tuple:
    page = browser.new_page(user_agent=UA)
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(wait_ms)
        return page.content(), None
    except Exception as exc:
        return None, str(exc)
    finally:
        page.close()


def _strip_emoji(text: str) -> str:
    return re.sub(r'^[^\w$£€\d]+', '', text).strip()


def _parse_dailyremote(soup, page_url: str) -> list[dict]:
    jobs = []
    for card in soup.select("article.card"):
        link = card.select_one("h2.job-position a")
        if not link:
            continue
        title = link.get_text(strip=True)
        href  = urljoin("https://dailyremote.com", link["href"])

        # Type + date from the two non-dot spans inside .company-name
        spans    = [s for s in card.select("div.company-name span") if s.get_text(strip=True) != "·"]
        job_type = spans[0].get_text(strip=True) if spans else "N/A"
        date     = spans[1].get_text(strip=True) if len(spans) > 1 else "N/A"

        # Salary, location, skill tags from card-tag badges
        salary, location = "Not specified", "N/A"
        skill_tags = []
        for tag in card.select("span.card-tag"):
            raw = tag.get_text(strip=True)
            clean = _strip_emoji(raw)
            if not clean:
                continue
            if "$" in raw or "£" in raw or "€" in raw or re.search(r'\bper\b', raw, re.I):
                salary = clean
            elif any(c in raw for c in ("🌎","🌍","🌏","🇺🇸")) or re.search(r'(United States|Remote|Worldwide)', raw, re.I):
                location = clean
            elif clean.lower() not in ("others", "other"):
                skill_tags.append(clean)

        # AI summary shown on the listing card itself
        summary_el = card.select_one(".ai-responsibilities")
        description = summary_el.get_text(strip=True) if summary_el else "N/A"

        jobs.append({
            "jobTitle":    title,
            "jobURL":      href,
            "location":    location,
            "datePosted":  date,
            "jobType":     job_type,
            "salary":      salary,
            "tags":        skill_tags,
            "description": description,
        })
    return jobs


def _parse_generic(soup, page_url: str) -> list[dict]:
    job_path_hints = ["/job", "/jobs/", "/position", "/career", "/remote-job", "/listing", "/opening", "/vacancy"]
    seen, jobs = set(), []
    for a in soup.find_all("a", href=True):
        href = urljoin(page_url, a["href"])
        if href in seen or not any(h in href.lower() for h in job_path_hints):
            continue
        title = a.get_text(strip=True)
        if not title or len(title) < 5:
            continue
        seen.add(href)
        container = a.find_parent(["article", "li", "tr"]) or a.find_parent("div")
        card_text = container.get_text(" ", strip=True) if container else ""
        sal_m = re.search(r'\$[\d,]+(?:\s*[-–]\s*\$?[\d,]+)?\s*(?:per\s+\w+|/\w+)?', card_text, re.I)
        salary = sal_m.group(0) if sal_m else "Not specified"
        jobs.append({
            "jobTitle":    title,
            "jobURL":      href,
            "location":    "N/A",
            "datePosted":  "N/A",
            "jobType":     "N/A",
            "salary":      salary,
            "tags":        [],
            "description": "N/A",
        })
        if len(jobs) >= 25:
            break
    return jobs


def scrape_listing_page(browser, url: str) -> tuple:
    html, err = fetch_html(browser, url, wait_ms=3000)
    if err or not html:
        return [], err
    soup = BeautifulSoup(html, "lxml")
    for tag in soup.select("script, style, nav, footer"):
        tag.decompose()
    hostname = urlparse(url).hostname or ""
    if "dailyremote" in hostname:
        return _parse_dailyremote(soup, url), None
    return _parse_generic(soup, url), None


def scrape_job_detail(browser, job_url: str, hostname: str) -> dict:
    if "dailyremote" in hostname:
        return {}
    html, _ = fetch_html(browser, job_url, wait_ms=2000)
    if not html:
        return {}
    soup = BeautifulSoup(html, "lxml")
    for tag in soup.select("script, style, nav, footer"):
        tag.decompose()

    company = "N/A"
    for sel in ["[class*='company']", "[class*='employer']", "[itemprop='hiringOrganization']"]:
        el = soup.select_one(sel)
        if el:
            t = el.get_text(strip=True)
            if t and len(t) < 80:
                company = t; break

    desc = "N/A"
    for sel in ["[class*='description']", "[class*='job-content']", "[class*='content']", "article", "main"]:
        el = soup.select_one(sel)
        if el and len(el.get_text(strip=True)) > 150:
            desc = el.get_text("\n", strip=True)[:3000]; break

    reqs = "N/A"
    for h in soup.find_all(["h2","h3","h4","strong"]):
        if any(kw in h.get_text().lower() for kw in ["require","qualif","must have","you need"]):
            parts = [s.get_text("\n", strip=True) for s in h.find_next_siblings(["p","ul","ol"])[:3]]
            if parts:
                reqs = "\n".join(parts); break

    apply_url = job_url
    for a in soup.find_all("a", href=True):
        if any(kw in a.get_text(strip=True).lower() for kw in ["apply now","apply here","submit"]):
            h = a["href"]
            apply_url = h if h.startswith("http") else urljoin(job_url, h); break

    return {"companyName": company, "jobDescription": desc, "requirements": reqs, "applyURL": apply_url}


def build_row(listing: dict, detail: dict) -> dict:
    tags_str = ", ".join(t for t in listing.get("tags", []) if t) or "N/A"
    desc = listing.get("description", "N/A")
    if detail.get("jobDescription") and detail["jobDescription"] != "N/A":
        desc = detail["jobDescription"]
    return {
        "Job Title":    listing.get("jobTitle", "N/A"),
        "Company":      detail.get("companyName", "N/A"),
        "Salary":       listing.get("salary", "Not specified"),
        "Location":     listing.get("location", "N/A"),
        "Job Type":     listing.get("jobType", "N/A"),
        "Date Posted":  listing.get("datePosted", "N/A"),
        "Tags":         tags_str,
        "Description":  desc,
        "Requirements": detail.get("requirements", "N/A"),
        "Apply URL":    detail.get("applyURL", listing.get("jobURL", "N/A")),
        "Listing URL":  listing.get("jobURL", "N/A"),
    }


# ── Excel export ───────────────────────────────────────────────────────────────

def to_excel_bytes(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    HEADER_FILL  = PatternFill(start_color="0C1E31", end_color="0C1E31", fill_type="solid")
    SALARY_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    SALARY_YELLOW= PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    SALARY_RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="D4AF37")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    salary_col = COLUMNS.index("Salary") + 1
    wrap_cols  = {"Description", "Requirements"}

    for row in rows:
        ws.append([row.get(c, "N/A") for c in COLUMNS])
        row_num = ws.max_row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = Alignment(wrap_text=(col_name in wrap_cols), vertical="top")
        sal_cell = ws.cell(row=row_num, column=salary_col)
        stype = salary_type(sal_cell.value or "")
        if stype == "disclosed":
            sal_cell.fill = SALARY_GREEN
        elif stype == "vague":
            sal_cell.fill = SALARY_YELLOW
        else:
            sal_cell.fill = SALARY_RED

    widths = {
        "Job Title": 38, "Company": 28, "Salary": 22, "Location": 20,
        "Job Type": 15, "Date Posted": 16, "Tags": 30,
        "Description": 60, "Requirements": 55, "Apply URL": 45, "Listing URL": 45,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Page config ────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Jerah's Job Scout",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(CSS, unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────

if "selected_board" not in st.session_state:
    st.session_state.selected_board = None
if "board_url" not in st.session_state:
    st.session_state.board_url = ""
if "active_terms" not in st.session_state:
    st.session_state.active_terms = set()
if "custom_url" not in st.session_state:
    st.session_state.custom_url = ""

# ── Sidebar ────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(
        '<h2 style="color:#D4AF37; margin-bottom:2px; font-size:20px;">Jerah\'s Job Scout</h2>'
        '<p style="color:#3d6080; font-size:12px; margin-top:0;">Salary-aware job scraping</p>',
        unsafe_allow_html=True,
    )
    st.divider()

    st.markdown(
        '<p style="color:#D4AF37; font-size:11px; font-weight:700; '
        'letter-spacing:1.5px; margin-bottom:8px;">SALARY COLOR KEY</p>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div style="background:#0c1e31; border:1px solid #162840; border-radius:10px; padding:14px;">'
        '<p style="margin:5px 0; font-size:13px; color:#4ecb78;">&#9679;&nbsp; Salary disclosed</p>'
        '<p style="margin:5px 0; font-size:13px; color:#f0cc55;">&#9679;&nbsp; Vague (e.g. Competitive)</p>'
        '<p style="margin:5px 0; font-size:13px; color:#ff7a7a;">&#9679;&nbsp; Not specified</p>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.divider()
    st.markdown('<p style="color:#253d54; font-size:11px;">Built with Playwright · Streamlit · openpyxl</p>', unsafe_allow_html=True)


# ── Header ─────────────────────────────────────────────────────────────────────

st.markdown(
    '<h1 style="color:#D4AF37; margin-bottom:2px; font-size:32px;">Jerah\'s Job Scout</h1>'
    '<p style="color:#3d6080; font-size:14px; margin-top:0; margin-bottom:24px;">'
    'Pick a board &nbsp;·&nbsp; Select your roles &nbsp;·&nbsp; Download your Excel</p>',
    unsafe_allow_html=True,
)

# ── Step 1: Board ──────────────────────────────────────────────────────────────

step_label(1, "Choose a job board")

board_cols = st.columns(len(JOB_BOARDS))
for i, (name, url) in enumerate(JOB_BOARDS.items()):
    with board_cols[i]:
        is_sel = st.session_state.selected_board == name
        label  = f"✓  {name}" if is_sel else name
        if st.button(label, key=f"board_{i}", use_container_width=True):
            st.session_state.selected_board = name
            st.session_state.board_url = url
            st.session_state.custom_url = ""   # clear any custom URL so it can't override
            st.rerun()

with st.expander("Paste a custom URL instead"):
    custom = st.text_input(
        "URL",
        key="custom_url",                      # bound to session_state — cleared by board buttons
        placeholder="https://your-job-board.com/jobs",
        label_visibility="collapsed",
    )
    if st.session_state.custom_url.strip():
        st.session_state.board_url = st.session_state.custom_url.strip()
        st.session_state.selected_board = "Custom"

if st.session_state.board_url:
    st.markdown(
        f'<p style="color:#3d6080; font-size:12px; margin-top:6px;">'
        f'&#9654;&nbsp; <span style="color:#D4AF37;">{st.session_state.board_url}</span></p>',
        unsafe_allow_html=True,
    )

st.markdown("<br>", unsafe_allow_html=True)

# ── Step 2: Terms ──────────────────────────────────────────────────────────────

step_label(2, "Select roles to search  (click to toggle)")

term_cols = st.columns(5)
for i, term in enumerate(QUICK_TERMS):
    with term_cols[i % 5]:
        active = term in st.session_state.active_terms
        label  = f"✓  {term}" if active else term
        if st.button(label, key=f"term_{i}", use_container_width=True):
            if active:
                st.session_state.active_terms.discard(term)
            else:
                st.session_state.active_terms.add(term)
            st.rerun()

extra_raw  = st.text_input(
    "Add more roles",
    placeholder="e.g. DevSecOps, Threat Intelligence Analyst  (comma-separated)",
)
extra_terms = [t.strip() for t in extra_raw.split(",") if t.strip()]
all_terms   = sorted(st.session_state.active_terms) + extra_terms

if all_terms:
    chips = "  ".join(
        f'<span style="background:#0c1e31; border:1px solid #D4AF37; color:#D4AF37; '
        f'border-radius:20px; padding:2px 10px; font-size:12px;">{t}</span>'
        for t in all_terms
    )
    st.markdown(f'<div style="margin-top:8px; line-height:2.2;">{chips}</div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── Step 3: Pages ──────────────────────────────────────────────────────────────

step_label(3, "How deep to search?")

max_pages = st.slider(
    "Pages per search term",
    min_value=1, max_value=20, value=5,
    help="Each page typically contains 10–25 listings. 5 pages is a good start.",
)

st.markdown("<br>", unsafe_allow_html=True)
st.divider()
st.markdown("<br>", unsafe_allow_html=True)

# ── Run button ─────────────────────────────────────────────────────────────────

run = st.button("Search Jobs", type="primary", use_container_width=True)

# ── Scrape logic ───────────────────────────────────────────────────────────────

if run:
    if not st.session_state.board_url:
        st.error("Select a job board first (Step 1).")
        st.stop()

    search_terms = all_terms or [""]
    seen_urls    = set()
    all_listings = []
    target_host  = urlparse(st.session_state.board_url).hostname or ""

    with browser_session() as browser:
        # ── Phase 1: collect listings ──────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        phase_label("Phase 1 — Collecting listings")
        p1_status = st.empty()
        p1_bar    = st.progress(0)

        scrape_errors = []
        total_terms   = len(search_terms)
        for t_idx, term in enumerate(search_terms):
            for page in range(1, max_pages + 1):
                url = build_url(st.session_state.board_url, term, page)
                status_line(
                    f'{gold(f"[{t_idx+1}/{total_terms}]")} &nbsp;'
                    f'"{term or "all"}" &nbsp;·&nbsp; page {page} &nbsp;·&nbsp; '
                    f'{white(str(len(all_listings)))} unique jobs found'
                    f'<br><span style="color:#3d5a73; font-size:11px;">{url}</span>',
                    p1_status,
                )
                listings, err = scrape_listing_page(browser, url)
                if err:
                    scrape_errors.append(f"Page {page} ({term or 'all'}): {err}")
                if not listings:
                    break
                added = 0
                for listing in listings:
                    job_url = listing.get("jobURL", "")
                    if job_url and job_url not in seen_urls:
                        seen_urls.add(job_url)
                        all_listings.append(listing)
                        added += 1
                if added == 0:
                    break
            p1_bar.progress((t_idx + 1) / total_terms)

        p1_status.success(f"Phase 1 complete — {len(all_listings)} unique jobs collected")

        if scrape_errors:
            with st.expander(f"Scrape warnings ({len(scrape_errors)})"):
                for e in scrape_errors:
                    st.warning(e)

        if not all_listings:
            st.error("No listings found. Check the URL and search terms.")
            st.stop()

        # ── Phase 2: scrape details ────────────────────────────────────────────
        phase_label("Phase 2 — Scraping job details")
        p2_status = st.empty()
        p2_bar    = st.progress(0)

        rows         = []
        table_holder = st.empty()
        total        = len(all_listings)
        start        = time.time()

        for i, listing in enumerate(all_listings, start=1):
            title   = listing.get("jobTitle", "Unknown")
            job_url = listing.get("jobURL", "")
            elapsed = time.time() - start
            avg     = elapsed / (i - 1) if i > 1 else 0
            eta     = f"{avg * (total - i + 1) / 60:.1f} min" if avg else "calculating..."

            status_line(
                f'{gold(f"[{i}/{total}]")} &nbsp;{title[:55]} &nbsp;·&nbsp; '
                f'ETA: {white(eta)}',
                p2_status,
            )
            p2_bar.progress(i / total)

            detail = scrape_job_detail(browser, job_url, target_host) if job_url else {}
            rows.append(build_row(listing, detail))

            if i % 5 == 0 or i == total:
                df_live = pd.DataFrame(rows)[
                    ["Job Title", "Company", "Salary", "Location", "Job Type", "Date Posted", "Apply URL"]
                ]
                table_holder.dataframe(df_live, use_container_width=True, height=280)

        p2_status.success(f"Done — all {total} jobs scraped")
        p2_bar.progress(1.0)

    # browser session closes here — results display doesn't need it

    # ── Results ────────────────────────────────────────────────────────────────
    st.divider()
    phase_label("Results")

    disclosed = sum(1 for r in rows if salary_type(r.get("Salary", "")) == "disclosed")
    vague     = sum(1 for r in rows if salary_type(r.get("Salary", "")) == "vague")
    missing   = total - disclosed - vague

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Jobs",         total)
    m2.metric("Salary Disclosed",   disclosed)
    m3.metric("Vague Salary",       vague)
    m4.metric("No Salary Listed",   missing)

    st.markdown("<br>", unsafe_allow_html=True)

    df_full = pd.DataFrame(rows)
    st.dataframe(
        df_full[["Job Title", "Company", "Salary", "Location", "Job Type", "Date Posted", "Apply URL"]],
        use_container_width=True,
        height=400,
    )

    st.divider()

    excel_bytes = to_excel_bytes(rows)
    st.download_button(
        label="Download Excel",
        data=excel_bytes,
        file_name="job_scout_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
