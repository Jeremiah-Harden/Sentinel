"""
Bulk-scrape multiple job searches from DailyRemote and export to Excel.

Usage:
    python tools/scrape_jobs_bulk.py --searches "cybersecurity intern,software engineer intern" --max-pages 20
    python tools/scrape_jobs_bulk.py --searches "security intern" --max-pages 5 --output .tmp/security_jobs.xlsx
"""

import argparse
import os
import sys
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from firecrawl import FirecrawlApp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()


LISTING_SCHEMA = {
    "type": "object",
    "properties": {
        "jobListings": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "jobTitle":   {"type": "string"},
                    "jobURL":     {"type": "string"},
                    "location":   {"type": "string"},
                    "datePosted": {"type": "string"},
                    "jobType":    {"type": "string"},
                    "salary":     {"type": "string"},
                    "tags":       {"type": "array", "items": {"type": "string"}},
                },
            },
        }
    },
}

DETAIL_SCHEMA = {
    "type": "object",
    "properties": {
        "companyName":      {"type": "string"},
        "jobDescription":   {"type": "string"},
        "requirements":     {"type": "string"},
        "applyURL":         {"type": "string"},
    },
}

COLUMNS = [
    "Job Title",
    "Company Name",
    "Location",
    "Job Type",
    "Salary",
    "Date Posted",
    "Tags / Skills",
    "Job Description",
    "Requirements",
    "Apply URL",
    "DailyRemote URL",
    "Search Term",
]

WIDTH = 70


def p(msg="", end="\n"):
    print(msg, end=end, flush=True)


def banner(search_terms, max_pages):
    p("=" * WIDTH)
    p("  TECH INTERNSHIP SCRAPER  |  DailyRemote  |  Powered by Firecrawl")
    p(f"  {len(search_terms)} search terms  ·  up to {max_pages} pages each  ·  full detail mode")
    p(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    p("=" * WIDTH)
    p()


def progress_bar(current, total, width=40, label=""):
    filled = int(width * current / total) if total else 0
    bar = "█" * filled + "░" * (width - filled)
    pct = 100 * current / total if total else 0
    p(f"  [{bar}] {current}/{total}  ({pct:.0f}%)  {label}", end="\r")


def _scrape_with_retry(app, url, options, retries=3, backoff=6.0):
    for attempt in range(1, retries + 1):
        try:
            result = app.v1.scrape_url(url, **options)
            raw = getattr(result, "json_field", None) or {}
            return raw if isinstance(raw, dict) else {}
        except Exception as exc:
            if attempt < retries:
                wait = backoff * attempt
                p(f"  ⚠  Attempt {attempt} failed ({exc.__class__.__name__}) — retrying in {wait:.0f}s...")
                time.sleep(wait)
            else:
                p(f"  ✗  Skipped after {retries} failed attempts: {url[:60]}")
                return {}


def scrape_listing_page(app, search, page):
    url = f"https://dailyremote.com/remote-jobs?search={search.replace(' ', '%20')}&page={page}"
    raw = _scrape_with_retry(
        app, url,
        {
            "formats": ["json"],
            "json_options": {
                "prompt": (
                    "Extract all job listings on this page. For each job include: "
                    "jobTitle, jobURL (full URL to the job detail page), location, "
                    "datePosted, jobType (internship/full-time/part-time/contract), "
                    "salary (or 'Not specified'), and tags (list of skill/category tags)."
                ),
                "schema": LISTING_SCHEMA,
            },
            "wait_for": 5000,
            "timeout": 120000,
        },
    )
    return raw.get("jobListings", []) if raw else []


def scrape_job_detail(app, job_url):
    return _scrape_with_retry(
        app, job_url,
        {
            "formats": ["json"],
            "json_options": {
                "prompt": (
                    "Extract the following fields from this job posting page: "
                    "companyName (the hiring company), jobDescription (full job description text), "
                    "requirements (qualifications and requirements section), "
                    "applyURL (the direct URL to apply, e.g. the 'Apply' button link — not the DailyRemote URL)."
                ),
                "schema": DETAIL_SCHEMA,
            },
            "wait_for": 3000,
            "timeout": 120000,
        },
    )


def build_row(listing, detail, search_term):
    tags = listing.get("tags", [])
    tags_str = ", ".join(t for t in tags if t and t.lower() != "others") or "N/A"
    return [
        listing.get("jobTitle", "N/A"),
        detail.get("companyName", "N/A"),
        listing.get("location", "N/A"),
        listing.get("jobType", "N/A"),
        listing.get("salary", "Not specified"),
        listing.get("datePosted", "N/A"),
        tags_str,
        detail.get("jobDescription", "N/A"),
        detail.get("requirements", "N/A"),
        detail.get("applyURL", "N/A"),
        listing.get("jobURL", "N/A"),
        search_term,
    ]


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30
    wrap_cols = {"Job Description", "Requirements"}

    for row in rows:
        ws.append(row)
        row_num = ws.max_row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_num, column=col_idx).alignment = Alignment(
                wrap_text=(col_name in wrap_cols),
                vertical="top",
            )

    max_widths = {
        "Job Title": 40, "Company Name": 30, "Location": 20,
        "Job Type": 15, "Salary": 25, "Date Posted": 18,
        "Tags / Skills": 30, "Job Description": 60, "Requirements": 60,
        "Apply URL": 50, "DailyRemote URL": 50, "Search Term": 30,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max_widths.get(col_name, 20)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Bulk-scrape DailyRemote jobs to Excel.")
    parser.add_argument("--searches", required=True,
                        help="Comma-separated search terms")
    parser.add_argument("--max-pages", type=int, default=10,
                        help="Max pages per search term (default: 10)")
    parser.add_argument("--output", default=".tmp/bulk_jobs.xlsx",
                        help="Output Excel path (default: .tmp/bulk_jobs.xlsx)")
    args = parser.parse_args()

    api_key = os.getenv("FIRECRAWL_API_KEY")
    if not api_key:
        p("Error: FIRECRAWL_API_KEY not found in .env", file=sys.stderr)
        sys.exit(1)

    app = FirecrawlApp(api_key=api_key)
    search_terms = [s.strip() for s in args.searches.split(",") if s.strip()]

    banner(search_terms, args.max_pages)

    # ── PHASE 1: collect all unique listings ─────────────────────────────
    p("┌" + "─" * (WIDTH - 2) + "┐")
    p("│  PHASE 1 — COLLECTING JOB LISTINGS" + " " * (WIDTH - 38) + "│")
    p("└" + "─" * (WIDTH - 2) + "┘")
    p()

    seen_urls: set[str] = set()
    all_listings: list[tuple[dict, str]] = []

    for term_idx, term in enumerate(search_terms, start=1):
        p(f"  [{term_idx:02d}/{len(search_terms):02d}]  Searching: \"{term}\"")
        term_new = 0
        for page in range(1, args.max_pages + 1):
            p(f"         → Page {page} ...", end="")
            listings = scrape_listing_page(app, term, page)
            if not listings:
                p(f"  (no results — end of search)")
                break
            new_count = 0
            for listing in listings:
                url = listing.get("jobURL", "")
                if url and url not in seen_urls:
                    seen_urls.add(url)
                    all_listings.append((listing, term))
                    new_count += 1
            term_new += new_count
            status = f"+{new_count} new" if new_count else "all duplicates"
            p(f"  {len(listings)} found  |  {status}  |  {len(all_listings)} total unique")
            if new_count == 0:
                p(f"         → Saturated — moving to next search term")
                break
        p(f"         ✓  '{term}' complete: {term_new} unique jobs added")
        p()

    p()
    p(f"  PHASE 1 COMPLETE — {len(all_listings)} unique jobs queued for detail scraping")
    p()

    if not all_listings:
        p("  No listings found. Check your search terms.")
        sys.exit(0)

    # ── PHASE 2: scrape job detail pages ─────────────────────────────────
    p("┌" + "─" * (WIDTH - 2) + "┐")
    p("│  PHASE 2 — SCRAPING JOB DETAIL PAGES" + " " * (WIDTH - 40) + "│")
    p("└" + "─" * (WIDTH - 2) + "┘")
    p()

    rows = []
    phase2_start = time.time()

    for i, (listing, term) in enumerate(all_listings, start=1):
        title = listing.get("jobTitle", "Unknown")
        job_url = listing.get("jobURL", "")

        elapsed = time.time() - phase2_start
        avg_per_job = elapsed / (i - 1) if i > 1 else 0
        remaining = avg_per_job * (len(all_listings) - i + 1)
        eta_str = f"{remaining / 60:.1f} min remaining" if avg_per_job else "calculating..."

        bar_width = 30
        filled = int(bar_width * (i - 1) / len(all_listings))
        bar = "█" * filled + "░" * (bar_width - filled)
        pct = 100 * (i - 1) / len(all_listings)

        p(f"  [{bar}] {i - 1}/{len(all_listings)}  {pct:.0f}%  |  {eta_str}")
        p(f"  [{i:03d}] {title[:60]}")
        p(f"        {job_url[:65]}")

        detail = scrape_job_detail(app, job_url) if job_url else {}
        company = detail.get("companyName", "N/A")
        p(f"        ✓  Company: {company[:50]}")
        p()

        rows.append(build_row(listing, detail, term))

    # Final bar
    p(f"  [{'█' * 30}] {len(all_listings)}/{len(all_listings)}  100%")
    p()

    # ── Write Excel ───────────────────────────────────────────────────────
    p("┌" + "─" * (WIDTH - 2) + "┐")
    p("│  PHASE 3 — WRITING EXCEL FILE" + " " * (WIDTH - 33) + "│")
    p("└" + "─" * (WIDTH - 2) + "┘")
    p()
    p(f"  Writing {len(rows)} rows to {args.output} ...")
    write_excel(rows, args.output)
    total_time = time.time() - phase2_start
    p(f"  ✓  Saved: {args.output}")
    p()
    p("=" * WIDTH)
    p(f"  DONE  |  {len(rows)} jobs  |  {total_time / 60:.1f} min total")
    p(f"  Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    p("=" * WIDTH)


if __name__ == "__main__":
    main()
