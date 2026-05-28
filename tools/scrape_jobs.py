"""
scrape_jobs.py — DailyRemote job scraper (Firecrawl-powered CLI tool).

Part of the WAT framework: this is a deterministic Tool layer script.
Claude (the Agent layer) calls this with specific arguments; it does
exactly one thing and returns a structured result (Excel file).

Two-phase scrape:
  Phase 1 — Listing page: extract job cards (title, URL, salary, tags, type).
  Phase 2 — Detail page:  visit each job URL to get company, description, requirements.

Why Firecrawl instead of requests + BeautifulSoup?
  DailyRemote renders some content client-side via JavaScript, which plain
  HTTP GET requests can't see. Firecrawl runs a real browser headlessly and
  also offers structured JSON extraction via LLM — so instead of writing brittle
  CSS selectors, we describe WHAT we want and the model extracts it.

Usage:
    python tools/scrape_jobs.py --search "cybersecurity intern" --page 1
    python tools/scrape_jobs.py --search "python developer" --page 2 --output .tmp/python_jobs.xlsx
"""

import argparse
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from firecrawl import FirecrawlApp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()


# JSON Schema objects tell Firecrawl's LLM exactly what fields to extract
# and what types to return. This is more reliable than CSS selectors because
# the model understands semantics — "jobTitle" finds the title even if the
# CSS class name changes, as long as it looks like a job title.
LISTING_SCHEMA = {
    "type": "object",
    "properties": {
        "jobListings": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "jobTitle":   {"type": "string"},
                    "jobURL":     {"type": "string"},
                    "location":   {"type": "string"},
                    "datePosted": {"type": "string"},
                    "jobType":    {"type": "string"},
                    "salary":     {"type": "string"},
                    "tags":       {"type": "array", "items": {"type": "string"}},
                },
            },
        }
    },
}

DETAIL_SCHEMA = {
    "type": "object",
    "properties": {
        "companyName":      {"type": "string"},
        "jobDescription":   {"type": "string"},
        "requirements":     {"type": "string"},
        "applyURL":         {"type": "string"},
    },
}

COLUMNS = [
    "Job Title",
    "Company Name",
    "Location",
    "Job Type",
    "Salary",
    "Date Posted",
    "Tags / Skills",
    "Job Description",
    "Requirements",
    "Apply URL",
    "DailyRemote URL",
]


def scrape_listing_page(app: FirecrawlApp, search: str, page: int) -> list[dict]:
    url = f"https://dailyremote.com/remote-jobs?search={search.replace(' ', '%20')}&page={page}"
    print(f"Scraping listing page: {url}")
    result = app.v1.scrape_url(
        url,
        formats=["json"],
        json_options={
            "prompt": (
                "Extract all job listings on this page. For each job include: "
                "jobTitle, jobURL (full URL to the job detail page), location, "
                "datePosted, jobType (internship/full-time/part-time/contract), "
                "salary (or 'Not specified'), and tags (list of skill/category tags)."
            ),
            "schema": LISTING_SCHEMA,
        },
        wait_for=5000,
    )
    raw = getattr(result, "json_field", None) or {}
    listings = raw.get("jobListings", []) if isinstance(raw, dict) else []
    print(f"Found {len(listings)} listings.")
    return listings


def scrape_job_detail(app: FirecrawlApp, job_url: str) -> dict:
    result = app.v1.scrape_url(
        job_url,
        formats=["json"],
        json_options={
            "prompt": (
                "Extract the following fields from this job posting page: "
                "companyName (the hiring company), jobDescription (full job description text), "
                "requirements (qualifications and requirements section), "
                "applyURL (the direct URL to apply, e.g. the 'Apply' button link — not the DailyRemote URL)."
            ),
            "schema": DETAIL_SCHEMA,
        },
        wait_for=3000,
    )
    raw = getattr(result, "json_field", None) or {}
    return raw if isinstance(raw, dict) else {}


def build_row(listing: dict, detail: dict) -> list:
    """Merge listing (card) data with detail (job page) data into one flat row.
    The "others" tag is filtered out — DailyRemote uses it as a catch-all
    category that adds noise without useful skill information.
    """
    tags = listing.get("tags", [])
    tags_str = ", ".join(t for t in tags if t and t.lower() != "others") or "N/A"
    return [
        listing.get("jobTitle", "N/A"),
        detail.get("companyName", "N/A"),
        listing.get("location", "N/A"),
        listing.get("jobType", "N/A"),
        listing.get("salary", "Not specified"),
        listing.get("datePosted", "N/A"),
        tags_str,
        detail.get("jobDescription", "N/A"),
        detail.get("requirements", "N/A"),
        detail.get("applyURL", "N/A"),
        listing.get("jobURL", "N/A"),
    ]


def write_excel(rows: list[list], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30

    wrap_cols = {"Job Description", "Requirements"}

    for row in rows:
        ws.append(row)
        row_num = ws.max_row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = Alignment(
                wrap_text=(col_name in wrap_cols),
                vertical="top",
            )

    # Auto-size columns with caps
    max_widths = {
        "Job Title": 40,
        "Company Name": 30,
        "Location": 20,
        "Job Type": 15,
        "Salary": 25,
        "Date Posted": 18,
        "Tags / Skills": 30,
        "Job Description": 60,
        "Requirements": 60,
        "Apply URL": 50,
        "DailyRemote URL": 50,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max_widths.get(col_name, 20)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nSaved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Scrape DailyRemote jobs to Excel.")
    parser.add_argument("--search", required=True, help="Search query (e.g. 'cybersecurity intern')")
    parser.add_argument("--page", type=int, default=1, help="Page number (default: 1)")
    parser.add_argument("--output", default=".tmp/jobs.xlsx", help="Output Excel path (default: .tmp/jobs.xlsx)")
    args = parser.parse_args()

    api_key = os.getenv("FIRECRAWL_API_KEY")
    if not api_key:
        print("Error: FIRECRAWL_API_KEY not found in .env", file=sys.stderr)
        sys.exit(1)

    app = FirecrawlApp(api_key=api_key)

    listings = scrape_listing_page(app, args.search, args.page)
    if not listings:
        print("No listings found. Check the search term or page number.")
        sys.exit(0)

    rows = []
    for i, listing in enumerate(listings, start=1):
        title = listing.get("jobTitle", "Unknown")
        job_url = listing.get("jobURL", "")
        print(f"Scraping job {i}/{len(listings)}: {title}")
        detail = scrape_job_detail(app, job_url) if job_url else {}
        rows.append(build_row(listing, detail))

    write_excel(rows, args.output)


if __name__ == "__main__":
    main()
